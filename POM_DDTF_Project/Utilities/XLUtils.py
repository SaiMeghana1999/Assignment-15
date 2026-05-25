import openpyxl

def get_row_count(file, sheet_name):

    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    rows = sheet.max_row
    workbook.close()
    return rows


def get_column_count(file, sheet_name):

    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    columns = sheet.max_column
    workbook.close()
    return columns


def read_data(file, sheet_name, row_num, column_num):

    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    data = sheet.cell(row=row_num,column=column_num).value

    workbook.close()

    return data


def write_data(file, sheet_name, row_num, column_num, data):

    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_num,column=column_num).value = data

    workbook.save(file)
    workbook.close()

